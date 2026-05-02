from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import SearchResult
from .normalize import map_access_type, map_degree, normalize_for_match, slugify
from .regions import infer_region


DEFAULT_CONSOLIDATED_FILE = "renati_resultados_consolidados.xlsx"

EXPORT_COLUMNS = [
    "id",
    "tema",
    "tipo_acceso",
    "palabras_clave",
    "fecha_publicacion",
    "enlace_documento_original",
    "resumen",
    "universidad",
    "region_inferida",
    "grado_tesis",
    "anio_publicacion",
    "mes_publicacion",
    "titulo",
    "autor",
    "renati_level_original",
    "renati_type_original",
]


def build_export_rows(
    items: list[SearchResult],
    topic: str,
    details: dict[str, object] | None = None,
) -> list[dict[str, object]]:
    topic_slug = slugify(topic)
    details = details or {}
    rows: list[dict[str, object]] = []
    for idx, item in enumerate(items, start=1):
        detail = details.get(item.document_url)
        detail_summary = getattr(detail, "summary", "") if detail else ""
        detail_keywords = getattr(detail, "keywords", "") if detail else ""
        detail_access_type = getattr(detail, "access_type", "") if detail else ""
        rows.append(
            {
                "id": f"RENATI-{topic_slug}-{idx:04d}",
                "tema": topic,
                "tipo_acceso": map_access_type(item.renati_type) or detail_access_type,
                "palabras_clave": detail_keywords or item.keywords,
                "fecha_publicacion": item.publication_date,
                "enlace_documento_original": item.document_url,
                "resumen": detail_summary,
                "universidad": item.university,
                "region_inferida": infer_region(item.university),
                "grado_tesis": map_degree(item.renati_level) or item.degree_name,
                "anio_publicacion": item.year,
                "mes_publicacion": item.month,
                "titulo": item.title,
                "autor": item.author,
                "renati_level_original": item.renati_level,
                "renati_type_original": item.renati_type,
            }
        )
    return rows


def export_excel(
    items: list[SearchResult],
    topic: str,
    output_dir: str | Path = "data/output",
    details: dict[str, object] | None = None,
    file_name: str = DEFAULT_CONSOLIDATED_FILE,
    append: bool = True,
) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    file_path = output_path / file_name
    rows = build_export_rows(items, topic=topic, details=details)
    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    if append and file_path.exists():
        previous = pd.read_excel(file_path, sheet_name="resultados")
        df = merge_consolidated_rows(previous, df)
    else:
        df = normalize_consolidated_ids(df)
    df.to_excel(file_path, index=False, sheet_name="resultados")
    _format_workbook(file_path)
    return file_path


def merge_consolidated_rows(previous: pd.DataFrame, current: pd.DataFrame) -> pd.DataFrame:
    previous = previous.reindex(columns=EXPORT_COLUMNS)
    current = current.reindex(columns=EXPORT_COLUMNS)
    merged = pd.concat([previous, current], ignore_index=True)
    if merged.empty:
        return merged
    merged["_has_summary"] = merged["resumen"].fillna("").astype(str).str.len() > 0
    merged["_summary_len"] = merged["resumen"].fillna("").astype(str).str.len()
    merged["_row_order"] = range(len(merged))
    merged["_dedupe_key"] = merged.apply(_dedupe_key, axis=1)
    merged = merged.sort_values(
        by=["_dedupe_key", "_has_summary", "_summary_len", "_row_order"],
        ascending=[True, True, True, True],
    )
    merged = merged.drop_duplicates(subset=["_dedupe_key"], keep="last")
    merged = merged.sort_values("_row_order").drop(
        columns=["_has_summary", "_summary_len", "_row_order", "_dedupe_key"]
    )
    merged["id"] = [f"RENATI-CONSOLIDADO-{idx:05d}" for idx in range(1, len(merged) + 1)]
    return merged[EXPORT_COLUMNS]


def normalize_consolidated_ids(df: pd.DataFrame) -> pd.DataFrame:
    df = df.reindex(columns=EXPORT_COLUMNS).copy()
    df = merge_consolidated_rows(pd.DataFrame(columns=EXPORT_COLUMNS), df)
    df["id"] = [f"RENATI-CONSOLIDADO-{idx:05d}" for idx in range(1, len(df) + 1)]
    return df[EXPORT_COLUMNS]


def _dedupe_key(row) -> str:
    title = normalize_for_match(str(row.get("titulo") or ""))
    university = normalize_for_match(str(row.get("universidad") or ""))
    if title:
        return f"title::{title}::university::{university}"
    return f"url::{normalize_for_match(str(row.get('enlace_documento_original') or ''))}"


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook["resultados"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 55)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    workbook.save(path)
